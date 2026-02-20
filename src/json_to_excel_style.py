# Utility to convert extracted JSON data into a styled Excel report
# Formats cells with colors, borders, and auto-adjusted widths

from __future__ import annotations

import argparse
import json
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, PatternFill, Border, Side

def collect_fields(obj, prefix=""):
    """Recursively collect all fields that have a 'value' key."""
    fields = []
    for k, v in obj.items():
        path = f"{prefix}.{k}" if prefix else k
        if path.startswith("meta."):
            continue
        if isinstance(v, dict) and "value" in v:
            label = path.split(".")[-1].replace("_", " ")
            fields.append((path, label))
        elif isinstance(v, dict):
            fields.extend(collect_fields(v, path))
    return fields


def extract_field(obj, path):
    parts = path.split(".")
    cursor = obj
    for part in parts:
        if not isinstance(cursor, dict) or part not in cursor:
            return None, None, None
        cursor = cursor[part]
    if not isinstance(cursor, dict):
        return cursor, None, None
    value = cursor.get("value")
    source = cursor.get("source") if isinstance(cursor.get("source"), dict) else {}
    page = source.get("page")
    excerpt = source.get("excerpt")
    return value, page, excerpt


def normalize_value(val):
    """Convert various data types (bool, list, dict) to string representation."""
    if val is None:
        return ""
    if isinstance(val, bool):
        return "Oui" if val else "Non"
    if isinstance(val, list):
        return ", ".join(normalize_value(v) for v in val)
    if isinstance(val, dict):
        return ", ".join(f"{k}: {v}" for k, v in val.items())
    return str(val)


def build_matrix(obj):
    fields = collect_fields(obj)
    labels, values, pages, excerpts = [], [], [], []
    for path, label in fields:
        val, page, excerpt = extract_field(obj, path)
        labels.append(label)
        values.append(normalize_value(val))
        pages.append("" if page is None else page)
        excerpts.append("" if excerpt is None else excerpt)
    return pd.DataFrame([labels, values, pages, excerpts])


def apply_styling(xlsx_path, df):
    """Apply visual styling: alternating row colors, borders, and column widths."""
    from openpyxl import load_workbook

    wb = load_workbook(xlsx_path)
    ws = wb.active

    fill_a = PatternFill("solid", fgColor="E2EFDA")  # vert clair
    fill_b = PatternFill("solid", fgColor="FFFFFF")  # blanc
    align = Alignment(wrap_text=True, vertical="top")
    thin = Side(style="thin", color="000000")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    max_col_widths = []

    for col_idx in range(1, df.shape[1] + 1):
        fill = fill_a if col_idx % 2 else fill_b
        col_width = 0
        for row_idx in range(1, 4 + 1):  # 4 lignes
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.fill = fill
            cell.alignment = align
            cell.border = border_all
            text = str(cell.value) if cell.value is not None else ""
            col_width = max(col_width, len(text))
        max_col_widths.append(min(col_width + 4, 80))

    for col_idx, width in enumerate(max_col_widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    ws.row_dimensions[1].height = 30
    ws.row_dimensions[2].height = 30
    ws.row_dimensions[3].height = 22
    ws.row_dimensions[4].height = 60

    wb.save(xlsx_path)


def main():
    parser = argparse.ArgumentParser(
        description="Convertit un JSON carnet d'entretien en Excel stylé."
    )
    parser.add_argument("input_json", type=Path, help="Chemin du fichier JSON source")
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="Chemin de sortie .xlsx (défaut: même nom que le JSON avec suffixe _styled)",
    )
    args = parser.parse_args()

    output = args.output or args.input_json.with_name(
        args.input_json.stem + "_styled.xlsx"
    )

    data = json.loads(args.input_json.read_text(encoding="utf-8"))
    df = build_matrix(data)
    output.parent.mkdir(parents=True, exist_ok=True)
    df.to_excel(output, index=False, header=False)
    apply_styling(output, df)
    print(f"Écrit : {output}")


if __name__ == "__main__":
    main()
